"""Excel dosyasından ürünleri toplu yükleme sayfası."""
import os
import logging
from pathlib import Path
from datetime import datetime

import openpyxl
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QFileDialog, QProgressBar, QTextEdit, QMessageBox, QMessageBox
)
from PyQt5.QtGui import QFont, QColor

logger = logging.getLogger(__name__)


class ExcelImportWorker(QThread):
    """Excel dosyasını arka planda işle."""
    progress = pyqtSignal(str)  # Log mesajı
    finished = pyqtSignal(dict)  # {success: bool, inserted: int, updated: int, errors: []}

    def __init__(self, db, file_path):
        super().__init__()
        self.db = db
        self.file_path = file_path
        self.errors = []

    def run(self):
        """Excel'i oku ve DB'ye yükle."""
        inserted = 0
        updated = 0

        try:
            self.progress.emit("📂 Excel dosyası açılıyor...")
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            ws = wb.active

            self.progress.emit(f"📋 Sheet: {ws.title}, Satır: {ws.max_row}")

            # Excel'den ürünleri oku (Row 5'ten başla, başlıklar Row 4)
            for row_idx in range(5, ws.max_row + 1):
                try:
                    barcode = ws.cell(row_idx, 2).value  # Sütun B
                    brand = ws.cell(row_idx, 3).value    # Sütun C
                    name = ws.cell(row_idx, 4).value     # Sütun D
                    retail_price = ws.cell(row_idx, 5).value  # Sütun E
                    dealer_price = ws.cell(row_idx, 6).value  # Sütun F

                    # Boş satırları atla
                    if not barcode or not name:
                        continue

                    # Türkçe büyüt
                    barcode = str(barcode).strip()
                    name = str(name).strip()
                    brand = str(brand).strip() if brand else ""

                    # Fiyatları float'a çevir
                    try:
                        retail_price = float(retail_price) if retail_price else 0
                        dealer_price = float(dealer_price) if dealer_price else 0
                    except (ValueError, TypeError):
                        retail_price = 0
                        dealer_price = 0

                    # MongoDB'de ara
                    existing = self.db["products"].find_one({"barcode": barcode})

                    if existing:
                        # Güncelle
                        self.db["products"].update_one(
                            {"barcode": barcode},
                            {"$set": {
                                "name": name,
                                "brand": brand,
                                "retail_price": retail_price,
                                "dealer_price": dealer_price,
                                "updated_at": datetime.now()
                            }}
                        )
                        updated += 1
                        self.progress.emit(f"  ✏️  Satır {row_idx}: {barcode} ({name}) GÜNCELLENDI")
                    else:
                        # Yeni ekle
                        self.db["products"].insert_one({
                            "barcode": barcode,
                            "name": name,
                            "brand": brand,
                            "retail_price": retail_price,
                            "dealer_price": dealer_price,
                            "stock": 0,
                            "created_at": datetime.now(),
                            "updated_at": datetime.now()
                        })
                        inserted += 1
                        self.progress.emit(f"  ✅ Satır {row_idx}: {barcode} ({name}) EKLENDİ")

                except Exception as e:
                    self.errors.append(f"Satır {row_idx}: {str(e)}")
                    self.progress.emit(f"  ❌ Satır {row_idx}: {str(e)}")

            self.progress.emit(f"\n✅ İçeri aktarım tamamlandı: {inserted} yeni, {updated} güncellendi")
            self.finished.emit({
                "success": True,
                "inserted": inserted,
                "updated": updated,
                "errors": self.errors
            })

        except Exception as e:
            msg = f"Excel okuma hatası: {str(e)}"
            logger.error(msg)
            self.progress.emit(f"❌ {msg}")
            self.finished.emit({
                "success": False,
                "inserted": 0,
                "updated": 0,
                "errors": [msg]
            })


class ExcelImportPage(QWidget):
    """Merkez'de Excel yükleme sayfası."""

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.worker = None
        self.setup_ui()

    def setup_ui(self):
        """Arayüz oluştur."""
        layout = QVBoxLayout()

        # Başlık
        title = QLabel("📊 Excel'den Ürün İçeri Aktar")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title.setFont(title_font)
        layout.addWidget(title)

        # Açıklama
        info = QLabel(
            "Excel dosyasından ürünleri DB'ye yükle. "
            "Barkodu olan ürünler güncellenir, yenileri eklenir.\n\n"
            "Beklenen format: Sütun B=Barkod, C=Marka, D=Adı, E=Perakende Fiyatı, F=Bayi Fiyatı"
        )
        info.setStyleSheet("color: #666; font-size: 11px;")
        layout.addWidget(info)

        # Dosya seçim butonu
        file_btn_layout = QHBoxLayout()
        self.file_label = QLabel("Dosya seçilmedi...")
        file_btn = QPushButton("📁 Excel Dosyası Seç")
        file_btn.clicked.connect(self._select_file)
        file_btn_layout.addWidget(self.file_label)
        file_btn_layout.addWidget(file_btn)
        layout.addLayout(file_btn_layout)

        # İçeri aktar butonu
        self.import_btn = QPushButton("📤 İçeri Aktar")
        self.import_btn.clicked.connect(self._import_excel)
        self.import_btn.setEnabled(False)
        self.import_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 8px;")
        layout.addWidget(self.import_btn)

        # İlerleme çubuğu
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)

        # Log kutusu
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Monospace", 9))
        self.log_text.setMinimumHeight(300)
        layout.addWidget(QLabel("📋 İşlem Günlüğü:"))
        layout.addWidget(self.log_text)

        self.setLayout(layout)
        self.selected_file = None

    def _select_file(self):
        """Dosya seç."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Excel Dosyası Seç",
            str(Path.home()),
            "Excel Files (*.xlsx *.xls);;Tüm Dosyalar (*)"
        )
        if file_path:
            self.selected_file = file_path
            self.file_label.setText(os.path.basename(file_path))
            self.import_btn.setEnabled(True)
            self.log_text.clear()

    def _import_excel(self):
        """Excel'i yükle."""
        if not self.selected_file:
            QMessageBox.warning(self, "Hata", "Lütfen Excel dosyası seçin!")
            return

        self.import_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.log_text.clear()

        self.worker = ExcelImportWorker(self.db, self.selected_file)
        self.worker.progress.connect(self._log_message)
        self.worker.finished.connect(self._on_import_finished)
        self.worker.start()

    def _log_message(self, msg: str):
        """Log mesajı ekle."""
        self.log_text.append(msg)

    def _on_import_finished(self, result: dict):
        """İçeri aktarım bittiğinde."""
        self.progress_bar.setVisible(False)
        self.import_btn.setEnabled(True)

        if result["success"]:
            summary = (
                f"✅ Başarılı!\n\n"
                f"Yeni ürünler: {result['inserted']}\n"
                f"Güncellenen: {result['updated']}\n"
                f"Toplam: {result['inserted'] + result['updated']}"
            )
            self.log_text.append(f"\n{summary}")
            QMessageBox.information(self, "Başarılı", summary)
        else:
            QMessageBox.critical(
                self,
                "Hata",
                f"İçeri aktarım başarısız:\n{result['errors'][0]}"
            )
